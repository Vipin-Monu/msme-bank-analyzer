from flask import Flask, render_template_string, request, send_file
import pdfplumber
import pandas as pd
import os
import re

app = Flask(__name__)

PASSWORD = "1234"   # simple password for now

HTML = """
<!DOCTYPE html>
<html>
<head>
<title>Bank Analyzer</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
</head>
<body style="font-family:Arial;text-align:center;padding:40px;">

<h2>Bank Statement Analyzer</h2>

<form method="POST" enctype="multipart/form-data">
<input type="password" name="password" placeholder="Enter Password" required><br><br>
<input type="file" name="file" accept=".pdf" required><br><br>
<button type="submit">Upload & Convert</button>
</form>

</body>
</html>
"""

@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        if request.form["password"] != PASSWORD:
            return "Wrong Password"

        file = request.files["file"]
        file.save("statement.pdf")

        data = []

        with pdfplumber.open("statement.pdf") as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    lines = text.split("\n")
                    for line in lines:
                        match = re.search(r'(\d{2}/\d{2}/\d{4}).*?(-?\d+\.\d{2})$', line)
                        if match:
                            date = match.group(1)
                            amount = match.group(2)
                            data.append([date, amount])

        df = pd.DataFrame(data, columns=["Date", "Amount"])
        file_name = "All_Transactions.xlsx"

df.to_excel(file_name, index=False)

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

wb = load_workbook(file_name)
ws = wb.active

# Header Bold
for cell in ws[1]:
    cell.font = Font(bold=True)

# Date format DD-MM-YYYY
for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
    for cell in row:
        cell.number_format = "DD-MM-YYYY"

# Auto column width
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(file_name)

        os.remove("statement.pdf")

        return send_file("All_Transactions.xlsx", as_attachment=True)

    return render_template_string(HTML)


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
    
